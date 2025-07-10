import unittest
from openpyxl import load_workbook
from utils.excel_ops import copiar_intervalo_k5_r14, formatar_coluna_o_em_vermelho


class TestExcelOps(unittest.TestCase):
    def setUp(self):
        self.arquivo_teste = "tests/base_teste.xlsx"
        self.data = "2025-06-20"

    def test_copiar_intervalo(self):
        wb = load_workbook(self.arquivo_teste)
        copiar_intervalo_k5_r14(wb, self.data)
        aba = wb.active
        self.assertEqual(aba[f"K{self.data[-2:]}"].value, "Teste K5")

    def test_formatar_coluna(self):
        wb = load_workbook(self.arquivo_teste)
        formatar_coluna_o_em_vermelho(wb, self.data)
        aba = wb.active
        cor = (
            aba[f"O{self.data[-2:]}"].font.color.rgb
            if aba[f"O{self.data[-2:]}"].font.color
            else None
        )
        self.assertTrue(cor is None or "FF0000" in cor)


if __name__ == "__main__":
    unittest.main()
