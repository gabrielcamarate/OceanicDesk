"""
Sistema de Cache para Operações Excel - OceanicDesk

IMPORTANTE: Este sistema mantém 100% da compatibilidade com as operações Excel existentes.
Todas as funções de leitura e escrita de planilhas continuam funcionando exatamente igual.

Este módulo adiciona:
1. Cache inteligente para operações de leitura pesadas
2. Invalidação automática baseada em modificação de arquivos
3. Cache de dados processados (DataFrames, valores extraídos)
4. Otimização de carregamento de workbooks
5. Integração com sistema de logging e métricas
"""

import os
import time
import hashlib
import pickle
from pathlib import Path
from typing import Any, Dict, Optional, Callable, Union, Tuple
from datetime import datetime, timedelta
from functools import wraps
import threading

# Import do sistema de logging (se disponível)
try:
    from utils.logger import log_operacao, log_performance, logger
    LOGGING_AVAILABLE = True
except ImportError:
    LOGGING_AVAILABLE = False

# Import do sistema de exceções (se disponível)
try:
    from utils.exceptions import FileOperationError, safe_execute
    EXCEPTIONS_AVAILABLE = True
except ImportError:
    EXCEPTIONS_AVAILABLE = False

# ============================================================================
# CONFIGURAÇÕES DO CACHE
# ============================================================================

class CacheConfig:
    """Configurações do sistema de cache"""
    
    # Diretório base do cache
    CACHE_DIR = Path.home() / ".oceanicdesk_cache"
    
    # Tempo de vida padrão do cache (em segundos)
    DEFAULT_TTL = 3600  # 1 hora
    
    # Tamanho máximo do cache em MB
    MAX_CACHE_SIZE_MB = 100
    
    # Extensões de arquivo que podem ser cacheadas
    CACHEABLE_EXTENSIONS = {".xlsx", ".xls", ".csv"}
    
    # Prefixos para diferentes tipos de cache
    WORKBOOK_PREFIX = "wb_"
    DATAFRAME_PREFIX = "df_"
    VALUE_PREFIX = "val_"
    PROCESSED_PREFIX = "proc_"


# ============================================================================
# SISTEMA DE CACHE PRINCIPAL
# ============================================================================

class ExcelCache:
    """
    Sistema de cache para operações Excel.
    Mantém compatibilidade total com operações existentes.
    """
    
    def __init__(self, cache_dir: Optional[Path] = None):
        self.cache_dir = cache_dir or CacheConfig.CACHE_DIR
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        self._memory_cache = {}
        
        # Log de inicialização
        if LOGGING_AVAILABLE:
            log_operacao("cache_init", "INICIADO", {
                "cache_dir": str(self.cache_dir),
                "max_size_mb": CacheConfig.MAX_CACHE_SIZE_MB
            })
    
    def _get_file_hash(self, file_path: Union[str, Path]) -> str:
        """Gera hash único baseado no caminho e timestamp do arquivo"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            return ""
        
        # Combina caminho, tamanho e timestamp de modificação
        stat = file_path.stat()
        content = f"{file_path.absolute()}_{stat.st_size}_{stat.st_mtime}"
        return hashlib.md5(content.encode()).hexdigest()
    
    def _get_cache_path(self, prefix: str, key: str) -> Path:
        """Gera caminho do arquivo de cache"""
        return self.cache_dir / f"{prefix}{key}.cache"
    
    def _is_cache_valid(self, cache_path: Path, ttl: int) -> bool:
        """Verifica se o cache ainda é válido"""
        if not cache_path.exists():
            return False
        
        # Verifica TTL
        cache_age = time.time() - cache_path.stat().st_mtime
        return cache_age < ttl
    
    def _save_to_cache(self, cache_path: Path, data: Any) -> bool:
        """Salva dados no cache"""
        try:
            with self._lock:
                with open(cache_path, 'wb') as f:
                    pickle.dump(data, f)
                
                if LOGGING_AVAILABLE:
                    log_operacao("cache_save", "SUCESSO", {
                        "cache_file": cache_path.name,
                        "size_bytes": cache_path.stat().st_size
                    })
                
                return True
        except Exception as e:
            if LOGGING_AVAILABLE:
                logger.error(f"Erro ao salvar cache {cache_path}: {e}")
            return False
    
    def _load_from_cache(self, cache_path: Path) -> Optional[Any]:
        """Carrega dados do cache"""
        try:
            with self._lock:
                with open(cache_path, 'rb') as f:
                    data = pickle.load(f)
                
                if LOGGING_AVAILABLE:
                    log_operacao("cache_load", "SUCESSO", {
                        "cache_file": cache_path.name,
                        "size_bytes": cache_path.stat().st_size
                    })
                
                return data
        except Exception as e:
            if LOGGING_AVAILABLE:
                logger.error(f"Erro ao carregar cache {cache_path}: {e}")
            return None
    
    def get_workbook_cache(self, file_path: Union[str, Path], 
                          data_only: bool = True, ttl: int = CacheConfig.DEFAULT_TTL) -> Optional[Any]:
        """
        Obtém workbook do cache ou None se não estiver disponível.
        Complementa load_workbook() existente sem substituí-lo.
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return None
        
        # Gera chave do cache
        file_hash = self._get_file_hash(file_path)
        cache_key = f"{file_hash}_{data_only}"
        cache_path = self._get_cache_path(CacheConfig.WORKBOOK_PREFIX, cache_key)
        
        # Verifica se cache é válido
        if self._is_cache_valid(cache_path, ttl):
            return self._load_from_cache(cache_path)
        
        return None
    
    def set_workbook_cache(self, file_path: Union[str, Path], workbook: Any, 
                          data_only: bool = True) -> bool:
        """
        Salva workbook no cache.
        Complementa operações de load_workbook() existentes.
        """
        file_path = Path(file_path)
        
        # Gera chave do cache
        file_hash = self._get_file_hash(file_path)
        cache_key = f"{file_hash}_{data_only}"
        cache_path = self._get_cache_path(CacheConfig.WORKBOOK_PREFIX, cache_key)
        
        return self._save_to_cache(cache_path, workbook)
    
    def get_dataframe_cache(self, file_path: Union[str, Path], sheet_name: Optional[str] = None,
                           skiprows: int = 0, ttl: int = CacheConfig.DEFAULT_TTL) -> Optional[Any]:
        """
        Obtém DataFrame do cache.
        Complementa pd.read_excel() existente sem substituí-lo.
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return None
        
        # Gera chave do cache
        file_hash = self._get_file_hash(file_path)
        cache_key = f"{file_hash}_{sheet_name}_{skiprows}"
        cache_path = self._get_cache_path(CacheConfig.DATAFRAME_PREFIX, cache_key)
        
        # Verifica se cache é válido
        if self._is_cache_valid(cache_path, ttl):
            return self._load_from_cache(cache_path)
        
        return None
    
    def set_dataframe_cache(self, file_path: Union[str, Path], dataframe: Any,
                           sheet_name: Optional[str] = None, skiprows: int = 0) -> bool:
        """
        Salva DataFrame no cache.
        Complementa operações de pd.read_excel() existentes.
        """
        file_path = Path(file_path)
        
        # Gera chave do cache
        file_hash = self._get_file_hash(file_path)
        cache_key = f"{file_hash}_{sheet_name}_{skiprows}"
        cache_path = self._get_cache_path(CacheConfig.DATAFRAME_PREFIX, cache_key)
        
        return self._save_to_cache(cache_path, dataframe)
    
    def get_value_cache(self, operation_key: str, ttl: int = CacheConfig.DEFAULT_TTL) -> Optional[Any]:
        """
        Obtém valor processado do cache.
        Para operações como buscar_valor_total_geral(), extrações, etc.
        """
        cache_path = self._get_cache_path(CacheConfig.VALUE_PREFIX, operation_key)
        
        # Verifica se cache é válido
        if self._is_cache_valid(cache_path, ttl):
            return self._load_from_cache(cache_path)
        
        return None
    
    def set_value_cache(self, operation_key: str, value: Any) -> bool:
        """
        Salva valor processado no cache.
        """
        cache_path = self._get_cache_path(CacheConfig.VALUE_PREFIX, operation_key)
        return self._save_to_cache(cache_path, value)
    
    def invalidate_file_cache(self, file_path: Union[str, Path]) -> int:
        """
        Invalida todos os caches relacionados a um arquivo específico.
        Útil quando arquivo é modificado.
        """
        file_path = Path(file_path)
        file_hash = self._get_file_hash(file_path)
        
        invalidated = 0
        
        # Remove caches relacionados ao arquivo
        for cache_file in self.cache_dir.glob(f"*{file_hash}*.cache"):
            try:
                cache_file.unlink()
                invalidated += 1
            except Exception as e:
                if LOGGING_AVAILABLE:
                    logger.warning(f"Erro ao invalidar cache {cache_file}: {e}")
        
        if LOGGING_AVAILABLE and invalidated > 0:
            log_operacao("cache_invalidate", "SUCESSO", {
                "file_path": str(file_path),
                "caches_removed": invalidated
            })
        
        return invalidated
    
    def clear_cache(self, older_than_hours: Optional[int] = None) -> int:
        """
        Limpa cache antigo ou todo o cache.
        """
        removed = 0
        cutoff_time = None
        
        if older_than_hours:
            cutoff_time = time.time() - (older_than_hours * 3600)
        
        for cache_file in self.cache_dir.glob("*.cache"):
            try:
                if cutoff_time is None or cache_file.stat().st_mtime < cutoff_time:
                    cache_file.unlink()
                    removed += 1
            except Exception as e:
                if LOGGING_AVAILABLE:
                    logger.warning(f"Erro ao remover cache {cache_file}: {e}")
        
        if LOGGING_AVAILABLE:
            log_operacao("cache_clear", "SUCESSO", {
                "files_removed": removed,
                "older_than_hours": older_than_hours
            })
        
        return removed
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """
        Retorna estatísticas do cache.
        """
        stats = {
            "cache_dir": str(self.cache_dir),
            "total_files": 0,
            "total_size_mb": 0.0,
            "by_type": {}
        }
        
        for cache_file in self.cache_dir.glob("*.cache"):
            try:
                size_bytes = cache_file.stat().st_size
                stats["total_files"] += 1
                stats["total_size_mb"] += size_bytes / (1024 * 1024)
                
                # Categoriza por tipo
                if cache_file.name.startswith(CacheConfig.WORKBOOK_PREFIX):
                    cache_type = "workbooks"
                elif cache_file.name.startswith(CacheConfig.DATAFRAME_PREFIX):
                    cache_type = "dataframes"
                elif cache_file.name.startswith(CacheConfig.VALUE_PREFIX):
                    cache_type = "values"
                else:
                    cache_type = "other"
                
                if cache_type not in stats["by_type"]:
                    stats["by_type"][cache_type] = {"count": 0, "size_mb": 0.0}
                
                stats["by_type"][cache_type]["count"] += 1
                stats["by_type"][cache_type]["size_mb"] += size_bytes / (1024 * 1024)
                
            except Exception:
                continue
        
        return stats


# ============================================================================
# INSTÂNCIA GLOBAL DO CACHE
# ============================================================================

# Cache global para uso em todo o sistema
excel_cache = ExcelCache()


# ============================================================================
# DECORATORS PARA CACHE AUTOMÁTICO
# ============================================================================

def cache_workbook(data_only: bool = True, ttl: int = CacheConfig.DEFAULT_TTL):
    """
    Decorator para cache automático de operações load_workbook().
    Mantém compatibilidade total com funções existentes.
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Tenta extrair caminho do arquivo dos argumentos
            file_path = None
            if args:
                file_path = args[0]
            elif 'filename' in kwargs:
                file_path = kwargs['filename']
            
            if not file_path:
                # Se não conseguir extrair caminho, executa função original
                return func(*args, **kwargs)
            
            # Tenta obter do cache
            start_time = time.time()
            cached_result = excel_cache.get_workbook_cache(file_path, data_only, ttl)
            
            if cached_result is not None:
                # Cache hit
                duration_ms = (time.time() - start_time) * 1000
                if LOGGING_AVAILABLE:
                    log_performance(f"cache_hit_{func.__name__}", duration_ms, {
                        "file_path": str(file_path),
                        "cache_type": "workbook"
                    })
                return cached_result
            
            # Cache miss - executa função original
            result = func(*args, **kwargs)
            
            # Salva no cache
            excel_cache.set_workbook_cache(file_path, result, data_only)
            
            # Log de performance
            duration_ms = (time.time() - start_time) * 1000
            if LOGGING_AVAILABLE:
                log_performance(f"cache_miss_{func.__name__}", duration_ms, {
                    "file_path": str(file_path),
                    "cache_type": "workbook"
                })
            
            return result
        
        return wrapper
    return decorator


def cache_dataframe(sheet_name: Optional[str] = None, skiprows: int = 0, 
                   ttl: int = CacheConfig.DEFAULT_TTL):
    """
    Decorator para cache automático de operações pd.read_excel().
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Tenta extrair parâmetros
            file_path = args[0] if args else kwargs.get('io')
            sheet = kwargs.get('sheet_name', sheet_name)
            skip = kwargs.get('skiprows', skiprows)
            
            if not file_path:
                return func(*args, **kwargs)
            
            # Tenta obter do cache
            start_time = time.time()
            cached_result = excel_cache.get_dataframe_cache(file_path, sheet, skip, ttl)
            
            if cached_result is not None:
                # Cache hit
                duration_ms = (time.time() - start_time) * 1000
                if LOGGING_AVAILABLE:
                    log_performance(f"cache_hit_{func.__name__}", duration_ms, {
                        "file_path": str(file_path),
                        "cache_type": "dataframe"
                    })
                return cached_result
            
            # Cache miss - executa função original
            result = func(*args, **kwargs)
            
            # Salva no cache
            excel_cache.set_dataframe_cache(file_path, result, sheet, skip)
            
            # Log de performance
            duration_ms = (time.time() - start_time) * 1000
            if LOGGING_AVAILABLE:
                log_performance(f"cache_miss_{func.__name__}", duration_ms, {
                    "file_path": str(file_path),
                    "cache_type": "dataframe"
                })
            
            return result
        
        return wrapper
    return decorator


def cache_operation(operation_key: str, ttl: int = CacheConfig.DEFAULT_TTL):
    """
    Decorator para cache de operações customizadas.
    Útil para funções como buscar_valor_total_geral(), extrações específicas, etc.
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Gera chave única baseada na função e argumentos
            args_str = str(args)[:100]  # Limita tamanho
            kwargs_str = str(kwargs)[:100]
            cache_key = f"{operation_key}_{hashlib.md5(f'{args_str}_{kwargs_str}'.encode()).hexdigest()}"

            # Tenta obter do cache
            start_time = time.time()
            cached_result = excel_cache.get_value_cache(cache_key, ttl)

            if cached_result is not None:
                # Cache hit
                duration_ms = (time.time() - start_time) * 1000
                if LOGGING_AVAILABLE:
                    log_performance(f"cache_hit_{func.__name__}", duration_ms, {
                        "operation_key": operation_key,
                        "cache_type": "operation"
                    })
                return cached_result

            # Cache miss - executa função original
            result = func(*args, **kwargs)

            # Salva no cache
            excel_cache.set_value_cache(cache_key, result)

            # Log de performance
            duration_ms = (time.time() - start_time) * 1000
            if LOGGING_AVAILABLE:
                log_performance(f"cache_miss_{func.__name__}", duration_ms, {
                    "operation_key": operation_key,
                    "cache_type": "operation"
                })

            return result

        return wrapper
    return decorator


# ============================================================================
# FUNÇÕES DE CONVENIÊNCIA PARA INTEGRAÇÃO
# ============================================================================

def cached_load_workbook(filename: Union[str, Path], data_only: bool = True,
                        ttl: int = CacheConfig.DEFAULT_TTL, **kwargs):
    """
    Versão com cache da função load_workbook().
    Complementa a função original sem substituí-la.
    """
    # Tenta obter do cache primeiro
    cached_wb = excel_cache.get_workbook_cache(filename, data_only, ttl)
    if cached_wb is not None:
        if LOGGING_AVAILABLE:
            log_operacao("cached_load_workbook", "CACHE_HIT", {
                "file_path": str(filename)
            })
        return cached_wb

    # Cache miss - carrega normalmente
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename, data_only=data_only, **kwargs)

        # Salva no cache
        excel_cache.set_workbook_cache(filename, wb, data_only)

        if LOGGING_AVAILABLE:
            log_operacao("cached_load_workbook", "CACHE_MISS", {
                "file_path": str(filename)
            })

        return wb
    except Exception as e:
        if EXCEPTIONS_AVAILABLE:
            raise FileOperationError(
                f"Erro ao carregar workbook: {e}",
                file_path=str(filename),
                operation="cached_load_workbook"
            )
        else:
            raise


def cached_read_excel(io: Union[str, Path], sheet_name: Optional[str] = None,
                     skiprows: int = 0, ttl: int = CacheConfig.DEFAULT_TTL, **kwargs):
    """
    Versão com cache da função pd.read_excel().
    Complementa a função original sem substituí-la.
    """
    # Tenta obter do cache primeiro
    cached_df = excel_cache.get_dataframe_cache(io, sheet_name, skiprows, ttl)
    if cached_df is not None:
        if LOGGING_AVAILABLE:
            log_operacao("cached_read_excel", "CACHE_HIT", {
                "file_path": str(io),
                "sheet_name": sheet_name
            })
        return cached_df

    # Cache miss - carrega normalmente
    try:
        import pandas as pd
        df = pd.read_excel(io, sheet_name=sheet_name, skiprows=skiprows, **kwargs)

        # Salva no cache
        excel_cache.set_dataframe_cache(io, df, sheet_name, skiprows)

        if LOGGING_AVAILABLE:
            log_operacao("cached_read_excel", "CACHE_MISS", {
                "file_path": str(io),
                "sheet_name": sheet_name
            })

        return df
    except Exception as e:
        if EXCEPTIONS_AVAILABLE:
            raise FileOperationError(
                f"Erro ao ler Excel: {e}",
                file_path=str(io),
                operation="cached_read_excel"
            )
        else:
            raise


def invalidate_cache_on_save(file_path: Union[str, Path]):
    """
    Invalida cache quando arquivo é salvo/modificado.
    Deve ser chamada após operações de escrita.
    """
    invalidated = excel_cache.invalidate_file_cache(file_path)

    if LOGGING_AVAILABLE and invalidated > 0:
        log_operacao("cache_invalidate_on_save", "SUCESSO", {
            "file_path": str(file_path),
            "caches_invalidated": invalidated
        })


# ============================================================================
# COMPATIBILIDADE COM SISTEMA EXISTENTE
# ============================================================================

# Aliases para manter compatibilidade
cache_excel_file = cached_load_workbook
cache_dataframe_file = cached_read_excel

# Função para verificar se cache está disponível
def is_cache_available() -> bool:
    """Verifica se sistema de cache está disponível"""
    return True

# Função para desabilitar cache temporariamente
_cache_enabled = True

def disable_cache():
    """Desabilita cache temporariamente"""
    global _cache_enabled
    _cache_enabled = False

def enable_cache():
    """Reabilita cache"""
    global _cache_enabled
    _cache_enabled = True

def is_cache_enabled() -> bool:
    """Verifica se cache está habilitado"""
    return _cache_enabled
