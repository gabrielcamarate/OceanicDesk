from config import (
    CAMINHO_PLANILHA,
    NOME_ABA,
    DATA_HOJE,
    DATA_ONTEM,
    LOGIN_SISTEMA,
    SENHA_SISTEMA,
)
from openpyxl import load_workbook
from tkinter import messagebox
import os

from utils.arquivos import criar_backup_planilha
from utils.excel_ops import (
    copiar_intervalo_k5_r14,
    formatar_coluna_o_em_vermelho,
    inserir_valor_planilha,
    inserir_litro_e_desconto_planilha,
    inserir_cashback_e_pix_planilha,
    inserir_litros_planilha,
    atualizando_planilhas_projecao
)
from utils.sistema import (
    auto_system_login,
    auto_system_relatorio_litro_e_desconto,
    auto_system_relatorio_cashback_e_pix,
    acessar_relatorio_subcategoria,
    relatorio_combustiveis,
    relatorio_bebida_nao_alcoolica,
    relatorio_bomboniere,
    relatorio_cerveja,
    relatorio_food,
    relatorio_isqueiros,
    relatorio_cigarro,
    posto_chacaltaya
)
from utils.email import enviar_relatorio
from interfaces.entrada_dados import coletar_litros_usuario
from interfaces.metodos_pagamento import coletar_formas_pagamento
from interfaces.valores_fechamento import abrir_janela_valores
from utils.automacao import extrair_valor_tmp, acessar_fechamento_caixa
from utils.automacao import acessar_fechamento_caixa, automatizar_fechamento_caixa
from datetime import datetime, timedelta
from dotenv import load_dotenv
from utils.alerta_visual import mostrar_alerta_visual, mostrar_alerta_progresso
load_dotenv()

# Global modificável se planilha for definida externamente
CAMINHO_PLANILHA_DINAMICO = None


def etapa1_backup_e_precos():
    mostrar_alerta_visual("Iniciando Etapa 1", "Backup e atualização de preços", tipo="info")
    
    # Validação de configuração
    if not CAMINHO_PLANILHA and not CAMINHO_PLANILHA_DINAMICO:
        mostrar_alerta_visual("Erro de Configuração", "Caminho da planilha não definido", tipo="error")
        raise EnvironmentError("CAMINHO_PLANILHA não definido no .env e nenhum caminho dinâmico selecionado.")
    
    caminho = CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA
    assert caminho is not None, "Caminho da planilha não pode ser None."
    
    mostrar_alerta_visual("Validando arquivo", f"Verificando: {os.path.basename(caminho)}", tipo="dev")
    
    if not os.path.exists(caminho):
        mostrar_alerta_visual("Arquivo não encontrado", f"Planilha não existe: {caminho}", tipo="error")
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")
    
    # Backup da planilha
    mostrar_alerta_visual("Criando backup", "Gerando cópia de segurança...", tipo="info")
    criar_backup_planilha(caminho)
    mostrar_alerta_visual("Backup concluído", "Cópia de segurança criada com sucesso", tipo="success")
    
    # Carregando planilha
    mostrar_alerta_visual("Carregando planilha", "Abrindo arquivo Excel...", tipo="dev")
    wb = load_workbook(caminho)
    mostrar_alerta_visual("Planilha carregada", "Arquivo aberto com sucesso", tipo="dev")
    
    # Copiando intervalo
    mostrar_alerta_visual("Copiando dados", "Transferindo preços K5:R14...", tipo="info")
    copiar_intervalo_k5_r14(wb, DATA_HOJE)
    
    # Formatando coluna
    mostrar_alerta_visual("Formatando dados", "Aplicando formatação vermelha...", tipo="dev")
    formatar_coluna_o_em_vermelho(wb, DATA_HOJE)
    
    # Salvando alterações
    mostrar_alerta_visual("Salvando alterações", "Persistindo modificações...", tipo="info")
    wb.save(caminho)
    
    mostrar_alerta_visual("Etapa 1 Concluída", "Backup e preços atualizados com sucesso!", tipo="success")
    messagebox.showinfo("Etapa 1", "Backup e preços atualizados com sucesso!")


def etapa2_minimercado():
    mostrar_alerta_visual("Iniciando Etapa 2", "Relatório Mini-Mercado", tipo="info")
    
    # Validação de credenciais
    if not LOGIN_SISTEMA or not SENHA_SISTEMA:
        mostrar_alerta_visual("Erro de Configuração", "Credenciais do sistema não definidas", tipo="error")
        raise EnvironmentError("LOGIN_SISTEMA ou SENHA_SISTEMA não definido no .env.")
    
    caminho = CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA
    assert caminho is not None, "Caminho da planilha não pode ser None."
    
    # Login no sistema
    mostrar_alerta_visual("Conectando ao sistema", "Realizando login...", tipo="info")
    auto_system_login(LOGIN_SISTEMA, SENHA_SISTEMA)
    mostrar_alerta_visual("Login realizado", "Conectado ao AutoSystem", tipo="success")
    
    # Extraindo valor
    mostrar_alerta_visual("Extraindo dados", "Processando relatório temporário...", tipo="info")
    valor = extrair_valor_tmp()
    mostrar_alerta_visual("Dados extraídos", f"Valor encontrado: R$ {valor:,.2f}", tipo="dev")
    
    # Inserindo na planilha
    mostrar_alerta_visual("Inserindo dados", "Salvando valor na planilha...", tipo="info")
    wb = load_workbook(caminho)
    inserir_valor_planilha(wb, valor, DATA_HOJE)
    wb.save(caminho)
    
    mostrar_alerta_visual("Etapa 2 Concluída", "Relatório mini-mercado processado!", tipo="success")
    messagebox.showinfo("Etapa 2", "Relatório mini-mercado salvo na planilha.")
    

def etapa3_litros_descontos():
    mostrar_alerta_visual("Iniciando Etapa 3", "Litros e Descontos", tipo="info")
    
    # Gerando relatório
    mostrar_alerta_visual("Gerando relatório", "Solicitando dados de litros e descontos...", tipo="info")
    auto_system_relatorio_litro_e_desconto()
    mostrar_alerta_visual("Relatório gerado", "Dados obtidos do sistema", tipo="success")
    
    # Inserindo na planilha
    mostrar_alerta_visual("Processando dados", "Inserindo litros e descontos...", tipo="info")
    inserir_litro_e_desconto_planilha(
        CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA, NOME_ABA
    )
    
    # Finalizando processo
    mostrar_alerta_visual("Finalizando processo", "Encerrando AutoSystem...", tipo="dev")
    os.system("taskkill /f /im AutoSystem.exe")
    
    mostrar_alerta_visual("Etapa 3 Concluída", "Litros e descontos inseridos com sucesso!", tipo="success")
    messagebox.showinfo("Etapa 3", "Litros e descontos inseridos com sucesso.")


def etapa4_cashback_pix():
    mostrar_alerta_visual("Iniciando Etapa 4", "Cashback e Pix", tipo="info")
    
    # Gerando relatório
    mostrar_alerta_visual("Gerando relatório", "Solicitando dados de cashback e pix...", tipo="info")
    cashback, pix_total = auto_system_relatorio_cashback_e_pix()
    mostrar_alerta_visual("Dados obtidos", f"Cashback: R$ {cashback:,.2f} | Pix: R$ {pix_total:,.2f}", tipo="dev")
    
    # Inserindo na planilha
    mostrar_alerta_visual("Inserindo valores", "Salvando cashback e pix na planilha...", tipo="info")
    inserir_cashback_e_pix_planilha(
        CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA, NOME_ABA, cashback, pix_total
    )
    
    mostrar_alerta_visual("Etapa 4 Concluída", "Cashback e pix processados com sucesso!", tipo="success")


def etapa5_insercao_litros():
    mostrar_alerta_visual("Iniciando Etapa 5", "Inserção Manual de Litros", tipo="info")
    
    # Coletando dados do usuário
    mostrar_alerta_visual("Aguardando entrada", "Abrindo interface de coleta...", tipo="info")
    litros = coletar_litros_usuario()
    mostrar_alerta_visual("Dados coletados", f"Litros inseridos: {litros}", tipo="dev")
    
    # Inserindo na planilha
    mostrar_alerta_visual("Salvando dados", "Inserindo litros na planilha...", tipo="info")
    inserir_litros_planilha(
        CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA, NOME_ABA, *litros
    )
    
    mostrar_alerta_visual("Etapa 5 Concluída", "Litros inseridos com sucesso!", tipo="success")


def etapa6_envio_email():
    mostrar_alerta_visual("Iniciando Etapa 6", "Envio de Relatório por E-mail", tipo="info")
    
    # Enviando relatório
    mostrar_alerta_visual("Enviando e-mail", "Preparando e enviando relatório...", tipo="info")
    enviar_relatorio(
        CAMINHO_PLANILHA_DINAMICO or CAMINHO_PLANILHA, DATA_ONTEM
    )
    
    mostrar_alerta_visual("Etapa 6 Concluída", "Relatório enviado por e-mail com sucesso!", tipo="success")
    messagebox.showinfo("Etapa 6", "Relatório enviado por e-mail!")


def etapa7_fechamento_caixa():
    mostrar_alerta_visual("Iniciando Etapa 7", "Fechamento de Caixa (EMSys)", tipo="info")
    
    # Acessando fechamento
    mostrar_alerta_visual("Acessando EMSys", "Conectando ao sistema de fechamento...", tipo="info")
    acessar_fechamento_caixa(DATA_ONTEM)
    mostrar_alerta_visual("EMSys acessado", "Sistema de fechamento aberto", tipo="success")
    
    # Automatizando processo
    mostrar_alerta_visual("Automatizando fechamento", "Executando automação de caixa...", tipo="info")
    automatizar_fechamento_caixa()
    
    mostrar_alerta_visual("Etapa 7 Concluída", "Fechamento de caixa finalizado!", tipo="success")


def etapa8_projecao_de_vendas() -> None:
    """
    Executa a Etapa 8, que inclui:
    - Atualização de planilhas projeção
    - Atualização completa de relatórios do Meu Controle
    """
    mostrar_alerta_visual("Iniciando Etapa 8", "Projeção de Vendas", tipo="info")
    
    hoje = datetime.now()
    dia_inicio = hoje.replace(day=1).strftime("%d/%m/%Y")
    dia_fim = (hoje - timedelta(days=1)).strftime("%d/%m/%Y")
    ontem = hoje - timedelta(days=1)
    ontem = ontem.day
    
    mostrar_alerta_visual("Configurando período", f"De: {dia_inicio} | Até: {dia_fim}", tipo="dev")
    
    # Processo comentado - descomente conforme necessário
    mostrar_alerta_visual("Atualizando projeções", "Processando planilhas de projeção...", tipo="info")
    # atualizando_planilhas_projecao()
    
    mostrar_alerta_visual("Acessando relatórios", "Conectando ao Meu Controle...", tipo="info")
    # acessar_relatorio_subcategoria()
    
    # Relatórios específicos
    relatorios = [
        ("Combustíveis", lambda: relatorio_combustiveis(hoje, dia_inicio, dia_fim, ontem)),
        ("Bebidas não alcoólicas", lambda: relatorio_bebida_nao_alcoolica(ontem)),
        ("Bomboniere", lambda: relatorio_bomboniere(ontem)),
        ("Cerveja", lambda: relatorio_cerveja(ontem)),
        ("Food", lambda: relatorio_food(ontem)),
        ("Cigarro", lambda: relatorio_cigarro(ontem)),
        ("Isqueiros", lambda: relatorio_isqueiros(ontem))
    ]
    
    for i, (nome, funcao) in enumerate(relatorios):
        progresso = int((i / len(relatorios)) * 100)
        mostrar_alerta_progresso(f"Processando {nome}", f"Relatório {i+1} de {len(relatorios)}", progresso)
        # funcao()  # Descomente conforme necessário
    
    # Posto Chacaltaya
    mostrar_alerta_visual("Processando Chacaltaya", "Atualizando dados do posto...", tipo="info")
    posto_chacaltaya()
    
    mostrar_alerta_visual("Etapa 8 Concluída", "Projeção de vendas finalizada com sucesso!", tipo="success")
    messagebox.showinfo("Etapa 8", "Relatórios atualizados com sucesso!")