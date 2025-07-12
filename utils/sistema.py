from utils.file_utils import aguardar_arquivo
from utils.duplicata_ocr import aguardar_usuario
import os
import time
from datetime import datetime, timedelta
import pyautogui
from utils.logger import logger
from tkinter import messagebox
from utils.excel_ops import extrair_valores_relatorio_combustivel_tmp, extrair_valores_relatorio_bebidas_nao_alcoolicas_tmp, relatorio_bomboniere_tmp, relatorio_cerveja_tmp, relatorio_isqueiro_tmp, relatorio_cigarro_tmp
from utils.relatorios.food import atualizar_meu_controle
from utils.extratores import salvar_planilha_emsys
from utils.file_utils import corrigir_cache_excel_com
from utils.helpers import esperar_elemento
from utils.path_utils import get_captura_path, get_system_path, get_desktop_path
from utils.alerta_visual import mostrar_alerta_visual, mostrar_alerta_progresso
import pygetwindow as gw
from dotenv import load_dotenv
load_dotenv()

def mostrar_area_de_trabalho():
    """Minimiza todas as janelas no Windows."""
    mostrar_alerta_visual("Minimizando janelas", "Preparando área de trabalho...", tipo="dev")
    
    import platform
    import ctypes
    import time

    if platform.system() == "Windows":
        ctypes.windll.user32.keybd_event(0x5B, 0, 0, 0)  # Win
        ctypes.windll.user32.keybd_event(0x44, 0, 0, 0)  # D
        ctypes.windll.user32.keybd_event(0x44, 0, 2, 0)  # D up
        ctypes.windll.user32.keybd_event(0x5B, 0, 2, 0)  # Win up
        time.sleep(1)
        mostrar_alerta_visual("Área de trabalho limpa", "Janelas minimizadas", tipo="success")


def abrir_auto_system():
    mostrar_alerta_visual("Abrindo AutoSystem", "Iniciando aplicação...", tipo="info")
    
    caminho_app = get_system_path("autosystem")
    
    if not os.path.exists(caminho_app):
        mostrar_alerta_visual("Erro: AutoSystem não encontrado", f"Caminho: {caminho_app}", tipo="error")
        raise FileNotFoundError(f"AutoSystem não encontrado em: {caminho_app}")
    
    mostrar_alerta_visual("Validando arquivo", f"AutoSystem encontrado: {os.path.basename(caminho_app)}", tipo="dev")
    
    mostrar_area_de_trabalho()
    os.startfile(caminho_app)
    logger.info("Abrindo AutoSystem...")
    
    mostrar_alerta_visual("Aguardando carregamento", "Aguardando tela de login...", tipo="info")
    esperar_elemento(get_captura_path("autosystem_login.png"), timeout=60)
    mostrar_alerta_visual("AutoSystem carregado", "Tela de login exibida", tipo="success")


def fazer_login(usuario: str, senha: str):
    mostrar_alerta_visual("Realizando login", f"Usuário: {usuario}", tipo="info")
    
    logger.info("Efetuando login...")
    pyautogui.write(usuario, interval=0.1)
    mostrar_alerta_visual("Usuário inserido", "Campo usuário preenchido", tipo="dev")
    
    pyautogui.press("tab")
    pyautogui.write(senha, interval=0.1)
    mostrar_alerta_visual("Senha inserida", "Campo senha preenchido", tipo="dev")
    
    pyautogui.press("enter")
    mostrar_alerta_visual("Login enviado", "Aguardando autenticação...", tipo="info")
    
    esperar_elemento(get_captura_path("autosystem_remove_alert.png"), timeout=60)
    mostrar_alerta_visual("Login realizado", "Autenticação bem-sucedida", tipo="success")
    
    pyautogui.moveTo(1095, 13)
    pyautogui.click()
    mostrar_alerta_visual("Alerta removido", "Interface limpa", tipo="dev")


def navegar_para_relatorio_produtividade():
    mostrar_alerta_visual("Navegando para relatório", "Acessando menu de produtividade...", tipo="info")
    
    logger.info("Navegando até relatório de produtividade...")
    time.sleep(2)
    pyautogui.moveTo(782, 452)
    pyautogui.click()
    mostrar_alerta_visual("Menu acessado", "Primeira seleção realizada", tipo="dev")
    
    time.sleep(2)
    pyautogui.moveTo(942, 330)
    pyautogui.click()
    mostrar_alerta_visual("Submenu selecionado", "Segunda seleção realizada", tipo="dev")
    
    time.sleep(2)
    pyautogui.moveTo(49, 31, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    pyautogui.moveTo(156, 268, duration=1)
    time.sleep(1)
    pyautogui.moveTo(342, 656, duration=0.8)
    pyautogui.click()
    time.sleep(2)
    
    mostrar_alerta_visual("Relatório acessado", "Tela de produtividade carregada", tipo="success")


def preencher_filtros_e_gerar():
    mostrar_alerta_visual("Preenchendo filtros", "Configurando parâmetros do relatório...", tipo="info")
    
    pyautogui.press("tab")
    pyautogui.write("1612", interval=0.1)
    mostrar_alerta_visual("Código inserido", "Código 1612 preenchido", tipo="dev")
    
    for _ in range(4):
        pyautogui.press("tab")
    pyautogui.write("50", interval=0.1)
    mostrar_alerta_visual("Quantidade inserida", "Quantidade 50 preenchida", tipo="dev")
    
    for _ in range(5):
        pyautogui.press("tab")

    data_ontem = (datetime.today() - timedelta(days=1)).strftime("%d/%m/%Y")
    pyautogui.write(data_ontem, interval=0.1)
    mostrar_alerta_visual("Data inicial", f"Data: {data_ontem}", tipo="dev")
    
    pyautogui.press("tab")
    pyautogui.write(data_ontem, interval=0.1)
    mostrar_alerta_visual("Data final", f"Data: {data_ontem}", tipo="dev")

    pyautogui.moveTo(465, 456, duration=0.8)
    pyautogui.click()
    time.sleep(1)
    mostrar_alerta_visual("Filtros aplicados", "Configurações salvas", tipo="dev")
    
    for _ in range(5):
        time.sleep(0.2)
        pyautogui.press("tab")
    pyautogui.press("space")
    for _ in range(6):
        pyautogui.press("tab")
    pyautogui.press("enter")
    time.sleep(2)
    
    mostrar_alerta_visual("Relatório gerado", "Dados processados com sucesso", tipo="success")


def exportar_relatorio_excel():
    mostrar_alerta_visual("Exportando relatório", "Preparando exportação para Excel...", tipo="info")
    
    logger.info("Exportando relatório para Excel...")
    pyautogui.moveTo(143, 39)
    pyautogui.click()
    time.sleep(1)
    pyautogui.moveTo(596, 451)
    pyautogui.click()
    time.sleep(6)
    mostrar_alerta_visual("Exportação iniciada", "Aguardando processamento...", tipo="dev")
    
    pyautogui.moveTo(194, 743, duration=0.5)
    time.sleep(0.5)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.press("f12")
    time.sleep(2)
    pyautogui.hotkey("alt", "d")
    time.sleep(0.5)
    pyautogui.write(get_desktop_path())
    pyautogui.press("enter")
    time.sleep(0.5)
    pyautogui.write("tmp")
    time.sleep(0.5)
    pyautogui.press("tab", interval=0.2)
    pyautogui.press("down", presses=1, interval=0.2)
    pyautogui.press("up", presses=28, interval=0.2)
    time.sleep(0.3)
    pyautogui.press("enter", presses=2, interval=0.2)
    pyautogui.hotkey("ctrl", "w")
    time.sleep(2)
    
    logger.info("✅ Relatório exportado.")
    mostrar_alerta_visual("Relatório exportado", "Arquivo tmp.xlsx salvo no Desktop", tipo="success")


def auto_system_login(usuario: str, senha: str):
    mostrar_alerta_visual("Iniciando AutoSystem", "Sequência completa de login", tipo="info")
    
    abrir_auto_system()
    fazer_login(usuario, senha)
    navegar_para_relatorio_produtividade()
    preencher_filtros_e_gerar()
    exportar_relatorio_excel()
    
    mostrar_alerta_visual("AutoSystem concluído", "Processo completo finalizado", tipo="success")


def abrir_relatorio_vendas_detalhado():
    mostrar_alerta_visual("Abrindo vendas detalhado", "Acessando relatório de vendas...", tipo="info")
    
    logger.info("Abrindo relatório de vendas detalhado...")
    pyautogui.click(109, 28)
    time.sleep(1)
    pyautogui.moveTo(250, 267, duration=0.5)
    time.sleep(1)
    pyautogui.moveTo(560, 108, duration=0.5)
    time.sleep(1)
    pyautogui.moveTo(771, 268, duration=0.5)
    time.sleep(1)
    pyautogui.click()
    
    mostrar_alerta_visual("Vendas detalhado aberto", "Relatório carregado com sucesso", tipo="success")


def aplicar_filtros_vendas_detalhado():
    mostrar_alerta_visual("Aplicando filtros", "Configurando parâmetros de vendas...", tipo="info")
    
    logger.info("Aplicando filtros...")
    pyautogui.click(1069, 212)
    time.sleep(1)
    pyautogui.click(888, 427)
    time.sleep(1)
    pyautogui.click(467, 350)
    time.sleep(1)
    pyautogui.click(644, 535)
    time.sleep(1)
    pyautogui.click(497, 374)
    time.sleep(0.5)
    pyautogui.press("backspace", presses=8, interval=0.2)
    data_ontem = (datetime.today() - timedelta(days=1)).strftime("%d/%m/%Y")
    pyautogui.write(data_ontem, interval=0.1)
    mostrar_alerta_visual("Data configurada", f"Período: {data_ontem}", tipo="dev")
    
    pyautogui.click(521, 471, duration=0.5)
    pyautogui.click(541, 531, duration=0.5)
    pyautogui.click(631, 530, duration=0.5)
    time.sleep(0.5)
    
    mostrar_alerta_visual("Filtros aplicados", "Configurações salvas", tipo="success")


def gerar_e_exportar_vendas_detalhado():
    mostrar_alerta_visual("Gerando vendas detalhado", "Processando relatório...", tipo="info")
    
    logger.info("Gerando e exportando relatório...")
    pyautogui.click(396, 678)
    time.sleep(4)
    mostrar_alerta_visual("Relatório processado", "Dados gerados com sucesso", tipo="dev")
    
    pyautogui.click(144, 38)
    time.sleep(0.5)
    pyautogui.click(593, 450)
    time.sleep(3)
    pyautogui.moveTo(194, 743, duration=0.5)
    time.sleep(1)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.press("f12")
    time.sleep(2)
    pyautogui.hotkey("alt", "d")
    time.sleep(0.5)
    pyautogui.write(get_desktop_path())
    pyautogui.press("enter")
    time.sleep(1.5)
    pyautogui.write("tmp")
    time.sleep(0.5)
    pyautogui.press("tab", interval=0.2)
    pyautogui.press("down", presses=1, interval=0.2)
    pyautogui.press("up", presses=28, interval=0.2)
    time.sleep(0.3)
    pyautogui.press("enter", presses=2, interval=0.2)
    pyautogui.hotkey("ctrl", "w")
    time.sleep(5)
    
    logger.info("✅ Relatório detalhado exportado.")
    mostrar_alerta_visual("Vendas detalhado exportado", "Arquivo salvo no Desktop", tipo="success")


def auto_system_relatorio_litro_e_desconto():
    abrir_relatorio_vendas_detalhado()
    aplicar_filtros_vendas_detalhado()
    gerar_e_exportar_vendas_detalhado()


def abrir_emsys3():
    logger.info("Abrindo EMSys3...")
    mostrar_area_de_trabalho()
    os.startfile(get_system_path("emsys3"))
    esperar_elemento(get_captura_path("emsys_login.png"), timeout=60)


def login_emsys3():
    mostrar_alerta_visual("Login EMSys3", "Iniciando login do usuário EMSys3...", tipo="info")
    pyautogui.press("down", presses=8, interval=0.2)
    pyautogui.press("enter")
    time.sleep(6.5)
    pyautogui.press("tab", presses=2, interval=0.2)
    usuario = os.getenv("USUARIO_NILTON")
    senha = os.getenv("SENHA_NILTON")
    if not usuario or not senha:
        raise EnvironmentError("Variáveis USUARIO_NILTON e SENHA_NILTON não definidas no .env")
    pyautogui.write(usuario)
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write(senha)
    pyautogui.press("enter")
    esperar_elemento(get_captura_path("emsys_alert.png"), timeout=60)
    mostrar_alerta_visual("Login realizado com sucesso", "Usuário autenticado no EMSys3.", tipo="success")


def acessar_relatorio_pix():
    pyautogui.click(1094, 134)
    time.sleep(1)
    pyautogui.click(155, 38)
    time.sleep(1)
    pyautogui.moveTo(529, 105, duration=0.4)
    pyautogui.click()
    time.sleep(0.8)
    pyautogui.moveTo(669, 142)
    pyautogui.click()
    time.sleep(2.5)
    pyautogui.click(901, 512)
    time.sleep(4.5)
    pyautogui.click(765, 482)
    time.sleep(6)


def gerar_relatorio_pix():
    data_ontem = (datetime.today() - timedelta(days=1)).strftime("%d/%m/%Y")
    pyautogui.write(data_ontem, interval=0.2)
    pyautogui.write(data_ontem, interval=0.2)
    pyautogui.press("down")
    time.sleep(0.8)
    pyautogui.press("tab")
    pyautogui.write("68", interval=0.1)
    pyautogui.press("tab", presses=2, interval=0.2)
    pyautogui.press("enter")
    time.sleep(2.5)
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)
    

def processar_relatorio_excel_cashback_pix():
    from openpyxl import load_workbook
    from openpyxl.cell import MergedCell
    import win32com.client as win32

    corrigir_cache_excel_com()

    desktop_tmp = os.path.join(os.path.expanduser("~"), "Desktop", "tmp.xlsx")
    aguardar_arquivo(desktop_tmp)
    excel = win32.gencache.EnsureDispatch("Excel.Application") # type: ignore
    excel.Visible = True
    wb = excel.Workbooks.Open(desktop_tmp)
    time.sleep(10)
    wb.Save()
    wb.Close(SaveChanges=True)
    excel.Quit()
    logger.info("🔄 Arquivo tmp.xlsx aberto e salvo via Excel COM.")

    wb_tmp = load_workbook(desktop_tmp, data_only=True)
    ws_tmp = wb_tmp.active
    cashback_valor = None
    pagarme_valor = None
    pix_valor = None

    if ws_tmp is not None:
        for row in ws_tmp.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    valor = str(cell.value).strip().upper()
                    if valor == "CASHBACK MARKA":
                        cashback_valor = row[15].value
                    elif valor == "PAGAR.ME INSTITUICAO DE PAGAMENTO S.A":
                        pagarme_valor = row[15].value
                    elif valor == "PIX - CIELO":
                        pix_valor = row[15].value
    else:
        logger.warning("ws_tmp é None. Não foi possível iterar pelas linhas.")

    def safe_float(val):
        from decimal import Decimal
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, Decimal):
            return float(val)
        try:
            return float(str(val).replace(',', '.'))
        except Exception:
            return 0.0

    pagarme_valor = pagarme_valor or 0
    pix_valor = pix_valor or 0
    total_pix = safe_float(pagarme_valor) + safe_float(pix_valor)

    if os.path.exists(desktop_tmp):
        os.remove(desktop_tmp)
        logger.info("🗑️ tmp.xlsx deletado.")

    return cashback_valor, total_pix



def auto_system_relatorio_cashback_e_pix():
    abrir_emsys3()
    login_emsys3()
    acessar_relatorio_pix()
    gerar_relatorio_pix()
    return processar_relatorio_excel_cashback_pix()


def acessar_relatorio_subcategoria():
    # Acessar menu de vendas > relatórios > venda por subcategoria
    pyautogui.click(239, 83)
    time.sleep(2)
    pyautogui.moveTo(381, 595)
    time.sleep(2)
    pyautogui.moveTo(698, 358, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    
    
def relatorio_combustiveis(hoje, dia_inicio, dia_fim, ontem, chacal=False): 
    pyautogui.press('tab')
    pyautogui.write('2')
    pyautogui.press('tab')
    pyautogui.write('7')
    pyautogui.press('tab')
    pyautogui.write('51')
    pyautogui.press('tab', presses=4)

    pyautogui.write(dia_inicio)
    time.sleep(0.5)
    pyautogui.write(dia_fim)

    # Confirmar geração do relatório
    pyautogui.moveTo(538, 440, duration=0.2)
    pyautogui.click()
    pyautogui.press('tab', presses=6)
    pyautogui.press('enter')
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)
    extrair_valores_relatorio_combustivel_tmp(ontem, chacal=chacal)
    

def relatorio_bebida_nao_alcoolica(ontem, chacal=False):
    time.sleep(2)
    pyautogui.press("tab", presses=4)
    pyautogui.write("1")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("4")
    pyautogui.press("tab", presses=13)
    time.sleep(0.5)
    pyautogui.press("enter")
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)    
    extrair_valores_relatorio_bebidas_nao_alcoolicas_tmp(ontem, chacal=chacal)
    

def relatorio_bomboniere(ontem, chacal=False):
    time.sleep(2)
    pyautogui.press("tab", presses=4)
    pyautogui.write("1")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("3")
    pyautogui.press("tab", presses=13)
    time.sleep(0.5)
    pyautogui.press("enter")
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)    
    relatorio_bomboniere_tmp(ontem, chacal=chacal)


def relatorio_cerveja(ontem, chacal=False):
    time.sleep(2)
    pyautogui.press("tab", presses=4)
    pyautogui.write("1")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("62")
    pyautogui.press("tab", presses=13)
    time.sleep(0.5)
    pyautogui.press("enter")
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)    
    relatorio_cerveja_tmp(ontem, chacal=chacal)


def relatorio_cigarro(ontem, chacal=False):
    time.sleep(2)
    pyautogui.press("tab", presses=4)
    pyautogui.write("1")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("1")
    pyautogui.press("tab", presses=13)
    time.sleep(0.5)
    pyautogui.press("enter")
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(657,140, duration=0.5)    
    relatorio_cigarro_tmp(ontem, chacal=chacal)

  
def relatorio_food(ontem, chacal=False):
    atualizar_meu_controle(dia_fim=ontem, chacal=chacal)


def relatorio_isqueiros(ontem, chacal=False):
    time.sleep(2)
    pyautogui.press("tab", presses=4)
    pyautogui.write("1")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("0")
    pyautogui.press("tab")
    time.sleep(0.5)
    pyautogui.write("538")
    pyautogui.press("tab", presses=12)
    time.sleep(0.5)
    pyautogui.press("enter")
    
    salvar_planilha_emsys()
    pyautogui.click(1334,42)
    pyautogui.click(829,146, duration=0.5)    
    relatorio_isqueiro_tmp(ontem, chacal=chacal)
    
    
def posto_chacaltaya():
    hoje = datetime.now()
    dia_inicio = hoje.replace(day=1).strftime("%d/%m/%Y")
    dia_fim = (hoje - timedelta(days=1)).strftime("%d/%m/%Y")
    ontem = hoje - timedelta(days=1)
    ontem = ontem.day
    
    pyautogui.click(83,44)    
    pyautogui.click(29,81, duration=0.2) 
    time.sleep(8)
    pyautogui.click(29,81, duration=0.2) 
    time.sleep(6)
    pyautogui.press("tab")
    usuario = os.getenv("USUARIO_ELIANE")
    senha = os.getenv("SENHA_ELIANE")
    if not usuario or not senha:
        raise EnvironmentError("Variáveis USUARIO_ELIANE e SENHA_ELIANE não definidas no .env")
    pyautogui.write(usuario)
    pyautogui.press("tab")
    pyautogui.write(senha)
    pyautogui.press("enter")
    time.sleep(30)
    pyautogui.click(155,38)
    time.sleep(3)
    acessar_relatorio_subcategoria()
    relatorio_combustiveis(hoje, dia_inicio, dia_fim, ontem, chacal=True)
    relatorio_bebida_nao_alcoolica(ontem, chacal=True)
    relatorio_bomboniere(ontem, chacal=True)
    relatorio_cerveja(ontem, chacal=True)
    relatorio_food(ontem, chacal=True)
    relatorio_cigarro(ontem, chacal=True)
    relatorio_isqueiros(ontem, chacal=True)