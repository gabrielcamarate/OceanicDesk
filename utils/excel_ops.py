from utils.file_utils import aguardar_arquivo
from openpyxl.styles import Font
from openpyxl.cell import MergedCell
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import xlwings as xw
import win32com.client as win32
import os
from pathlib import Path
from dotenv import load_dotenv
from projecao.vendas import atualizar_projecao_vendas
from projecao.combustiveis import (
    atualizar_combustiveis,
    atualizar_dados_projecao_combustiveis,
)
from projecao.consolidado import atualizar_valores_de_vendas_geral
from tkinter import messagebox
from interfaces.alerta_visual import mostrar_alerta_visual, mostrar_alerta_progresso
import time
import calendar
from datetime import datetime
load_dotenv()

hoje = datetime.today()
# dias_do_mes = calendar.monthrange(hoje.year, hoje.month)[1]
dias_do_mes = 30
LETRA_PLANILHA = "G"

def copiar_intervalo_k5_r14(wb, data_referencia):
    dia_origem = data_referencia - timedelta(days=2)
    dia_destino = data_referencia - timedelta(days=1)
    nome_aba_origem = f"Dia {dia_origem.day:02d}"
    nome_aba_destino = f"Dia {dia_destino.day:02d}"

    if nome_aba_origem not in wb.sheetnames or nome_aba_destino not in wb.sheetnames:
        mostrar_alerta_visual("Erro: Abas não encontradas", f"Verificar: {nome_aba_origem} ou {nome_aba_destino}", tipo="error")
        raise ValueError(
            f"Aba '{nome_aba_origem}' ou '{nome_aba_destino}' não encontrada no arquivo."
        )

    sheet_origem = wb[nome_aba_origem]
    sheet_destino = wb[nome_aba_destino]
    
    # Contador para progresso
    total_celulas = 0
    celulas_copiadas = 0
    
    for row in range(5, 15):
        for col in range(11, 19):
            total_celulas += 1
            valor = sheet_origem.cell(row=row, column=col).value
            destino_cell = sheet_destino.cell(row=row, column=col)
            if isinstance(destino_cell, MergedCell):
                continue
            destino_cell.value = valor
            celulas_copiadas += 1
    print(
        f"[Cópia] Intervalo K5:R14 copiado de '{nome_aba_origem}' para '{nome_aba_destino}'."
    )

def formatar_coluna_o_em_vermelho(wb, data_referencia):
    
    dia_anterior = data_referencia - timedelta(days=1)
    nome_aba = f"Dia {dia_anterior.day:02d}"
    
    
    if nome_aba not in wb.sheetnames:
        mostrar_alerta_visual("Erro: Aba não encontrada", f"Aba '{nome_aba}' não existe", tipo="error")
        raise ValueError(f"Aba '{nome_aba}' não encontrada no arquivo.")

    aba = wb[nome_aba]
    celulas_formatadas = 0
    
    for row in range(5, 15):
        cell = aba.cell(row=row, column=15)
        if cell.value is not None:
            cell.font = Font(color="FF0000")
            celulas_formatadas += 1
    
    print(
        f"[Formato] Células O5:O14 formatadas com fonte vermelha na aba '{nome_aba}'."
    )
    

def inserir_valor_planilha(wb, valor, data_referencia):
    
    nome_aba = f"Dia {(data_referencia - timedelta(days=1)).day:02d}"
    
    
    if nome_aba in wb.sheetnames:
        aba = wb[nome_aba]
        aba["D33"] = valor
        print(f"[Planilha] Valor {valor} inserido na aba '{nome_aba}' célula D33.")
    else:
        mostrar_alerta_visual("Erro: Aba não encontrada", f"Aba '{nome_aba}' não existe", tipo="error")
        raise ValueError(f"Aba '{nome_aba}' não encontrada no arquivo.")


def inserir_litro_e_desconto_planilha(caminho_arquivo, nome_aba):
    mostrar_alerta_visual("Inserindo litros e descontos", "Processando dados temporários...", tipo="info")
    
    import os
    import time
    import pandas as pd
    from openpyxl import load_workbook

    desktop_tmp = os.path.join(os.path.expanduser("~"), "Desktop", "tmp.xlsx")
    aguardar_arquivo(desktop_tmp)

    # Espera até 30 segundos para o arquivo existir
    tempo_maximo = 30  # segundos
    intervalo = 0.5  # segundos
    tempo_esperado = 0

    while not os.path.exists(desktop_tmp) and tempo_esperado < tempo_maximo:
        time.sleep(intervalo)
        tempo_esperado += intervalo
        if tempo_esperado % 5 == 0:  # Alerta a cada 5 segundos
            mostrar_alerta_visual("Aguardando arquivo", f"Tempo: {tempo_esperado}s / {tempo_maximo}s", tipo="warning")

    if not os.path.exists(desktop_tmp):
        mostrar_alerta_visual("Erro: Arquivo não encontrado", f"tmp.xlsx não apareceu em {tempo_maximo}s", tipo="error")
        raise FileNotFoundError(
            f"Arquivo {desktop_tmp} não encontrado após {tempo_maximo} segundos."
        )

    # Processa a planilha temporária
    df = pd.read_excel(desktop_tmp, engine="openpyxl", skiprows=9)
    df = df[df["Produtos"].str.strip().str.upper() != "TOTAL"]

    dados = {}
    produtos_processados = 0
        
    for _, row in df.iterrows():
        produto = row["Produtos"].strip().upper()
        quantidade = row["Quantidade"] if pd.notna(row["Quantidade"]) else 0
        desconto = row["Desconto"] if pd.notna(row["Desconto"]) else 0
        dados[produto] = {"Litragem": quantidade, "Desconto": desconto}
        produtos_processados += 1

    mostrar_alerta_visual("Dados processados", f"{produtos_processados} produtos extraídos", tipo="dev")

    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]
    mapa_celulas = {
        "GASOLINA COMUM": {"quantidade": "D21", "desconto": "F21"},
        "GASOLINA ADITIVADA": {"quantidade": "D22", "desconto": "F22"},
        "ETANOL COMUM": {"quantidade": "D25", "desconto": "F25"},
        "OLEO DIESEL B S10 ADITIVADO": {"quantidade": "D28", "desconto": "F28"},
    }

    for produto, celulas in mapa_celulas.items():
        dados_produto = dados.get(produto.upper(), {"Litragem": 0, "Desconto": 0})
        ws[celulas["quantidade"]] = dados_produto["Litragem"]
        ws[celulas["desconto"]] = dados_produto["Desconto"]
    
    wb.save(caminho_arquivo)
    print("✅ Quantidade e Desconto inseridos com sucesso.")

    if os.path.exists(desktop_tmp):
        os.remove(desktop_tmp)
        print("🗑️ Arquivo tmp.xlsx deletado.")

    mostrar_alerta_visual("Litros e descontos inseridos", "Dados salvos com sucesso!", tipo="success")


def inserir_cashback_e_pix_planilha(
    caminho_arquivo, nome_aba, cashback_valor, total_pix
):
    mostrar_alerta_visual("Inserindo cashback e pix", "Salvando valores na planilha...", tipo="info")
    
    wb_destino = load_workbook(caminho_arquivo)
    ws_destino = wb_destino[nome_aba]
    
    if cashback_valor is not None:
        ws_destino.cell(row=21, column=8).value = cashback_valor
        mostrar_alerta_visual("Cashback inserido", f"Valor: R$ {cashback_valor:,.2f}", tipo="dev")
    else:
        mostrar_alerta_visual("Cashback não encontrado", "Valor nulo ou vazio", tipo="warning")
        print("⚠️ Valor do cashback não encontrado.")

    ws_destino.cell(row=21, column=13).value = total_pix
    mostrar_alerta_visual("Pix inserido", f"Total: R$ {total_pix:,.2f}", tipo="dev")
    
    wb_destino.save(caminho_arquivo)
    print("✅ Valores inseridos com sucesso.")
    mostrar_alerta_visual("Cashback e pix salvos", "Valores inseridos com sucesso!", tipo="success")


def inserir_litros_planilha(
    caminho_arquivo, nome_aba, gasolina_comum, gasolina_aditivada, etanol, diesel
):
    mostrar_alerta_visual("Inserindo litros", "Salvando dados de combustíveis...", tipo="info")
    
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]
    
    litros_inseridos = 0
    
    if etanol is not None:
        ws.cell(row=45, column=5).value = etanol
        mostrar_alerta_visual("Etanol inserido", f"Valor: {etanol}L", tipo="dev", tempo=1000)
        time.sleep(1)
        litros_inseridos += 1
        
    if gasolina_aditivada is not None:
        ws.cell(row=42, column=6).value = gasolina_aditivada
        mostrar_alerta_visual("Gasolina aditivada inserida", f"Valor: {gasolina_aditivada}L", tipo="dev", tempo=1000)
        time.sleep(1)
        litros_inseridos += 1
        
    if gasolina_comum is not None:
        ws.cell(row=41, column=10).value = gasolina_comum
        mostrar_alerta_visual("Gasolina comum inserida", f"Valor: {gasolina_comum}L", tipo="dev", tempo=1000)
        time.sleep(1)
        litros_inseridos += 1
        
    if diesel is not None:
        ws.cell(row=48, column=8).value = diesel
        mostrar_alerta_visual("Diesel inserido", f"Valor: {diesel}L", tipo="dev", tempo=1000)
        time.sleep(1)
        litros_inseridos += 1

    wb.save(caminho_arquivo)
    print("✅ Litros inseridos na planilha com sucesso.")
    mostrar_alerta_visual("Litros inseridos", f"{litros_inseridos} valores salvos com sucesso!", tipo="success")


def atualizando_planilhas_projecao():
    mostrar_alerta_visual("Atualizando projeções", "Processando planilhas de projeção...", tipo="info")
    
    caminho_chacaltaya = os.getenv("CAMINHO_CHACALTAYA")
    caminho_oceanico_vendas = os.getenv("CAMINHO_PLANILHA_VENDAS")
    caminho_oceanico = os.getenv("CAMINHO_PLANILHA")
    caminho_controle = os.getenv("CAMINHO_MEU_CONTROLE")
    caminho_combustivel = os.getenv("CAMINHO_COMBUSTIVEL")
    caminho_copia = os.getenv("CAMINHO_COPIA_MES")

    caminhos = [
            caminho_chacaltaya,
            caminho_oceanico_vendas,
            caminho_oceanico,
            caminho_controle,
            caminho_combustivel,
            caminho_copia,
        ]
    nomes = [
        "CAMINHO_CHACALTAYA",
        "CAMINHO_PLANILHA_VENDAS",
        "CAMINHO_PLANILHA",
        "CAMINHO_MEU_CONTROLE",
        "CAMINHO_COMBUSTIVEL",
        "CAMINHO_COPIA_MES",
    ]
    
    
    faltantes = [n for n, v in zip(nomes, caminhos) if not v]
    if faltantes:
        mostrar_alerta_visual("Erro de Configuração", f"Variáveis ausentes: {', '.join(faltantes)}", tipo="error")
        raise EnvironmentError(f"Variável(is) de caminho ausente(s) no .env: {', '.join(faltantes)}")

    
    atualizar_projecao_vendas(
        caminho_arquivo_chacaltaya=caminho_chacaltaya,
        caminho_arquivo_vendas=caminho_oceanico_vendas,
        caminho_arquivo_destino=caminho_controle,
    )

    atualizar_combustiveis(
        caminho_vendas=caminho_oceanico,
        caminho_destino=caminho_combustivel,
    )

    atualizar_dados_projecao_combustiveis(
        caminho_vendas=caminho_oceanico,
        caminho_projecao=caminho_copia,
    )

    atualizar_valores_de_vendas_geral(
        caminho_arquivo_copia=caminho_copia,
        caminho_arquivo_meu_controle=caminho_controle,
    )
    
    return messagebox.showinfo("Etapa 8", "Valores atualizados nas planilhas.")
    

def extrair_valores_relatorio_combustivel_tmp(ontem, chacal=False):
    """
    Lê os valores do relatório tmp.xlsx e insere projeções como fórmulas
    na planilha 'Meu Controle' no caminho especificado no .env.
    
    Fórmula usada: =valor/dia_fim*30
    """
    load_dotenv()
    
    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Carrega valores do relatório temporário
    wb_tmp = load_workbook(caminho_tmp, data_only=True)
    ws_tmp = wb_tmp.active

    valores = {
        "gasolina_comum": ws_tmp["I14"].value,
        "gasolina_aditivada": ws_tmp["I20"].value,
        "etanol_comum": ws_tmp["I11"].value,
        "diesel_s10": ws_tmp["I17"].value,
    }

    # Carrega planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active  # ou nome da aba, se necessário

    # Mapeamento: célula destino -> variável
    if chacal:    
        destino = {
            f"{LETRA_PLANILHA}10": "gasolina_comum",
            f"{LETRA_PLANILHA}11": "gasolina_aditivada",
            f"{LETRA_PLANILHA}12": "etanol_comum",
            f"{LETRA_PLANILHA}13": "diesel_s10",
        }
    else:
        destino = {
            f"{LETRA_PLANILHA}32": "gasolina_comum",
            f"{LETRA_PLANILHA}33": "gasolina_aditivada",
            f"{LETRA_PLANILHA}34": "etanol_comum",
            f"{LETRA_PLANILHA}36": "diesel_s10",
        }

    # Inserir fórmulas
    for celula, nome_var in destino.items():
        valor = valores[nome_var]
        if isinstance(valor, (int, float)):
            formula = f"={valor}/{ontem}*{dias_do_mes}"
            ws_controle[celula] = formula

    wb_controle.save(caminho_meu_controle)
    
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")
    

def buscar_valor_total_geral(path_planilha: str, chacal=False) -> float:
    """
    Procura pela linha onde está o texto 'Total Geral (Todos os Departamentos)' na Coluna A
    e retorna o valor da Coluna R três linhas acima.
    """
    wb = load_workbook(path_planilha, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and "Total Geral (Todos os Departamentos)" in cell.value:
            linha = cell.row
            if chacal:
                valor = ws[f"K{linha}"].value  
                print(f"[Valor encontrado] K{linha} = {valor}")
                return valor
            else:    
                valor = ws[f"R{linha}"].value  
                print(f"[Valor encontrado] R{linha} = {valor}")
                return valor

    raise ValueError("Texto 'Total Geral (Todos os Departamentos)' não encontrado na Coluna A.")


def extrair_valores_relatorio_bebidas_nao_alcoolicas_tmp(dia_fim: int, chacal=False):
    """
    Extrai valor do tmp.xlsx e insere fórmula de projeção em Meu Controle na célula G39.
    A fórmula será: =valor/dia_fim*30
    """
    load_dotenv()

    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Extrair valor da planilha temporária
    valor = buscar_valor_total_geral(caminho_tmp, chacal=chacal)

    try:
        if isinstance(valor, str):
            valor = float(valor.replace(".", "").replace(",", "."))
    except ValueError:
        raise ValueError(f"Não foi possível converter o valor '{valor}' para float.")


    # Formatar valor com vírgula decimal (para Excel em português)
    valor_str = f"{valor:.2f}"  # mantém ponto decimal (ex: "9991.68")

    # Atualizar planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active

    formula = f"={valor_str}/{dia_fim}*{dias_do_mes}"
    
    if chacal:
        ws_controle[f"{LETRA_PLANILHA}17"] = formula
    else:
        ws_controle[f"{LETRA_PLANILHA}39"] = formula

    wb_controle.save(caminho_meu_controle)

    # Remover tmp.xlsx
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")


def relatorio_bomboniere_tmp(dia_fim: int, chacal=False):
    """
    Extrai valor do tmp.xlsx e insere fórmula de projeção em Meu Controle na célula G39.
    A fórmula será: =valor/dia_fim*30
    """
    load_dotenv()

    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Extrair valor da planilha temporária
    valor = buscar_valor_total_geral(caminho_tmp, chacal=chacal)

    try:
        if isinstance(valor, str):
            valor = float(valor.replace(".", "").replace(",", "."))
    except ValueError:
        raise ValueError(f"Não foi possível converter o valor '{valor}' para float.")


    # Formatar valor com vírgula decimal (para Excel em português)
    valor_str = f"{valor:.2f}"  # mantém ponto decimal (ex: "9991.68")

    # Atualizar planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active

    formula = f"={valor_str}/{dia_fim}*{dias_do_mes}"
    if chacal:
        ws_controle[f"{LETRA_PLANILHA}18"] = formula
    else:
        ws_controle[f"{LETRA_PLANILHA}40"] = formula
    wb_controle.save(caminho_meu_controle)

    # Remover tmp.xlsx
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")

    
def relatorio_cerveja_tmp(dia_fim: int, chacal=False):
    """
    Extrai valor do tmp.xlsx e insere fórmula de projeção em Meu Controle na célula G39.
    A fórmula será: =valor/dia_fim*30
    """
    load_dotenv()

    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Extrair valor da planilha temporária
    valor = buscar_valor_total_geral(caminho_tmp, chacal=chacal)

    try:
        if isinstance(valor, str):
            valor = float(valor.replace(".", "").replace(",", "."))
    except ValueError:
        raise ValueError(f"Não foi possível converter o valor '{valor}' para float.")


    # Formatar valor com vírgula decimal (para Excel em português)
    valor_str = f"{valor:.2f}"  # mantém ponto decimal (ex: "9991.68")

    # Atualizar planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active

    formula = f"={valor_str}/{dia_fim}*{dias_do_mes}"
    if chacal:
        ws_controle[f"{LETRA_PLANILHA}19"] = formula
    else:
        ws_controle[f"{LETRA_PLANILHA}41"] = formula
    wb_controle.save(caminho_meu_controle)

    # Remover tmp.xlsx
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")


def relatorio_cigarro_tmp(dia_fim: int, chacal=False):
    """
    Extrai valor do tmp.xlsx e insere fórmula de projeção em Meu Controle na célula G39.
    A fórmula será: =valor/dia_fim*30
    """
    load_dotenv()

    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Extrair valor da planilha temporária
    valor = buscar_valor_total_geral(caminho_tmp, chacal=chacal)

    try:
        if isinstance(valor, str):
            valor = float(valor.replace(".", "").replace(",", "."))
    except ValueError:
        raise ValueError(f"Não foi possível converter o valor '{valor}' para float.")


    # Formatar valor com vírgula decimal (para Excel em português)
    valor_str = f"{valor:.2f}"  # mantém ponto decimal (ex: "9991.68")

    # Atualizar planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active

    formula = f"={valor_str}/{dia_fim}*{dias_do_mes}"
    if chacal:
        ws_controle[f"{LETRA_PLANILHA}22"] = formula
    else:
        ws_controle[f"{LETRA_PLANILHA}44"] = formula
    wb_controle.save(caminho_meu_controle)

    # Remover tmp.xlsx
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")
    
    
def relatorio_isqueiro_tmp(dia_fim: int, chacal=False):
    """
    Extrai valor do tmp.xlsx e insere fórmula de projeção em Meu Controle na célula G39.
    A fórmula será: =valor/dia_fim*30
    """
    load_dotenv()

    # Caminhos
    caminho_tmp = Path.home() / "Desktop" / "tmp.xlsx"
    caminho_meu_controle = os.getenv("CAMINHO_MEU_CONTROLE")

    # Extrair valor da planilha temporária
    valor = buscar_isqueiro(caminho_tmp, chacal=chacal)

    # Formatar valor com vírgula decimal (para Excel em português)
    valor_str = f"{valor}"  # mantém ponto decimal (ex: "9991.68")

    # Atualizar planilha Meu Controle
    wb_controle = load_workbook(caminho_meu_controle)
    ws_controle = wb_controle.active

    formula = f"={valor_str}/{dia_fim}*{dias_do_mes}"
    if chacal:
        ws_controle[f"{LETRA_PLANILHA}21"] = formula
    else:
        ws_controle[f"{LETRA_PLANILHA}43"] = formula
    wb_controle.save(caminho_meu_controle)

    # Remover tmp.xlsx
    os.remove(caminho_tmp)
    print(f"[Limpeza] Arquivo temporário removido: {caminho_tmp}")


def buscar_isqueiro(caminho_tmp, chacal=False):
                
    if chacal:
        nomes_procurados = [
            "ISQUEIRO CLIPPER MINI SPECIAL",						
            "ISQUEIRO BIC MINI",		
            "ISQUEIRO BIC MAXXI",						
            "ISQUEIRO BIC MAXXI TREND MUSIC"
        ]
    else:		
        nomes_procurados = [
            "ISQUEIRO BIC MINI",
            "ISQUEIRO BIC MAXXI",
            "ISQUEIRO CRICKET MINI",
            "ISQUEIRO CLIPPER MAXI LISO",
            "ISQUEIRO ZENGAZ EMBORRACHADO GRAND JET CORES"
        ]

    wb = load_workbook(caminho_tmp, data_only=True)
    ws = wb.active

    valores = []

    for row in ws.iter_rows(min_row=1):
        nome = str(row[1].value).strip().upper() if row[1].value else ""
        if nome in nomes_procurados:
            valor_celula = row[8].value  # Coluna I = índice 8
            if isinstance(valor_celula, str):
                valor_celula = valor_celula.replace(".", "").replace(",", ".")
            try:
                valor = float(valor_celula)
                valores.append(valor)
            except (TypeError, ValueError):
                continue

    total = round(sum(valores), 2)
    print(f"Total de isqueiros vendidos: {total}")
    return total