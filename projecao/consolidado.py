import os
import openpyxl
from openpyxl import load_workbook
from pathlib import Path


def atualizar_valores_de_vendas_geral(
    caminho_arquivo_copia: str, caminho_arquivo_meu_controle: str
) -> None:
    from win32com.client import gencache

    caminho_arquivo_copia = Path(caminho_arquivo_copia)
    caminho_arquivo_meu_controle = Path(caminho_arquivo_meu_controle)

    if not caminho_arquivo_copia.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo_copia}")

    excel = gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False

    try:
        wb_com = excel.Workbooks.Open(caminho_arquivo_copia)
        wb_com.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        wb_com.Save()
        wb_com.Close()
    finally:
        excel.Quit()

    wb = load_workbook(caminho_arquivo_copia, data_only=True)
    ws = wb["VENDAS GERAL"]
    posto = ws.cell(row=35, column=13).value
    minimercado = ws.cell(row=35, column=15).value

    wb_controle = openpyxl.load_workbook(caminho_arquivo_meu_controle)
    aba = wb_controle.active
    aba["H27"] = posto
    aba["H31"] = minimercado
    wb_controle.save(caminho_arquivo_meu_controle)

    print(f"Arquivo modificado salvo como {caminho_arquivo_meu_controle}")
